#!/usr/bin/env python3
"""Portable launcher, first-use bootstrap, and fixed UI protocol for H."""

from __future__ import annotations

import argparse
import csv
import getpass
import hashlib
import json
import os
import platform
import shutil
import subprocess
import sys
import tarfile
import tempfile
import time
import urllib.request
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from fastmoss_client import (
    FastMossError,
    normalize_pids,
    query_products,
    save_product_results,
    validate_local_reference_image,
)


IS_FROZEN = bool(getattr(sys, "frozen", False))
PLUGIN_ROOT = (
    Path(sys.executable).resolve().parent.parent
    if IS_FROZEN
    else Path(__file__).resolve().parents[1]
)
REQUIREMENTS = PLUGIN_ROOT / "requirements.txt"
MAIN_SCRIPT = PLUGIN_ROOT / "scripts" / "kie_video_batch.py"
CODEX_HOME = Path(os.environ.get("CODEX_HOME", str(Path.home() / ".codex"))).expanduser()
CACHE_ROOT = CODEX_HOME / "cache" / "h"
USER_KEY_FILE = CODEX_HOME / "secrets" / "h_kie_api_key.txt"
FASTMOSS_KEY_FILE = CODEX_HOME / "secrets" / "h_fastmoss_api_key.txt"
LOCAL_KEY_FILE = PLUGIN_ROOT / ".h_api_key"
REQUIRED_IMPORTS = ["requests", "openpyxl"]
MIN_PYTHON = (3, 10)
PROTOCOL_VERSION = "h-fixed-v3"
GREETING = "哈喽小杨，你又开始工作啦，想不想小黄啊？"
MODE_MENU = "请选择功能，回复编号即可：\n1. PID\n2. 生成\n3. 发布"
GENERATE_MENU = "请选择生成方式，回复编号即可：\n1. 批处理\n2. 单处理"
ADSPOWER_RUNTIME_ARCHIVE = PLUGIN_ROOT / "assets" / "adspower-runtime.zip"
ADSPOWER_RUNTIME_SHA256 = "f1b5622348d5e632abb19d8af041f73e4523944fc9c0450d181dfe2ce0cef78c"
ADSPOWER_RUNTIME_ID = f"bundle-{ADSPOWER_RUNTIME_SHA256[:16]}"
ADSPOWER_RUNTIME_DIR = CACHE_ROOT / "adspower-runtime" / ADSPOWER_RUNTIME_ID
ADSPOWER_CLI = ADSPOWER_RUNTIME_DIR / "src" / "cli.mjs"
ADSPOWER_CONFIG_TEMPLATE = ADSPOWER_RUNTIME_DIR / "config.adspower.tiktok.example.json"
ADSPOWER_SCHEDULE_TEMPLATE = PLUGIN_ROOT / "assets" / "adspower-schedule-template.xlsx"
ADSPOWER_SCHEDULE_CSV = PLUGIN_ROOT / "assets" / "adspower-schedule.csv"
NODE_VERSION = "v24.18.1"
NODE_ASSETS = {
    "windows-x64": {
        "filename": f"node-{NODE_VERSION}-win-x64.zip",
        "sha256": "ec56b84a7551893ab2324ebdfdc4ab974a63b4781162600b68a1293cc3e53765",
        "executable": "node.exe",
    },
    "macos-intel": {
        "filename": f"node-{NODE_VERSION}-darwin-x64.tar.gz",
        "sha256": "6fb20fceacbb157c2f95825b80df4a454a0f6d81cdcd7bb81eeae9147e0e76ec",
        "executable": "bin/node",
    },
    "macos-apple-silicon": {
        "filename": f"node-{NODE_VERSION}-darwin-arm64.tar.gz",
        "sha256": "eb02f7fab96d3d67de40c5ec8566096fcb4c2026728787683ae5a97eb612b941",
        "executable": "bin/node",
    },
}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".webm"}


def configure_utf8_output() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass


def print_json(value: dict[str, Any]) -> None:
    configure_utf8_output()
    print(json.dumps(value, ensure_ascii=False, indent=2), flush=True)


def requirements_hash() -> str:
    payload = REQUIREMENTS.read_bytes() if REQUIREMENTS.exists() else b""
    payload += f"|py{sys.version_info.major}.{sys.version_info.minor}".encode("ascii")
    return hashlib.sha256(payload).hexdigest()[:16]


RUNTIME_ID = requirements_hash()
VENV_DIR = CACHE_ROOT / "venvs" / RUNTIME_ID
READY_FILE = CACHE_ROOT / f"ready-{RUNTIME_ID}.json"
DEPENDENCY_MARKER = VENV_DIR / ".h-requirements.sha256"
LOCK_FILE = CACHE_ROOT / "bootstrap.lock"


def packaged_core_path() -> Path | None:
    executable_name = "h_core.exe" if os.name == "nt" else "h_core"
    candidates = []
    if IS_FROZEN:
        candidates.append(Path(sys.executable).resolve().parent / executable_name)
    candidates.append(PLUGIN_ROOT / "runtime" / executable_name)
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


@dataclass(frozen=True)
class BootstrapReport:
    python: str
    environment_created: bool
    dependencies_installed: bool
    missing_before: list[str]
    marker_was_current: bool
    runtime_source: str


def venv_python() -> Path:
    if os.name == "nt":
        return VENV_DIR / "Scripts" / "python.exe"
    return VENV_DIR / "bin" / "python"


def run_quiet(
    command: list[str],
    *,
    cwd: Path | None = None,
    timeout: int = 600,
    env_overrides: dict[str, str] | None = None,
) -> subprocess.CompletedProcess[str]:
    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    if env_overrides:
        env.update(env_overrides)
    try:
        return subprocess.run(
            command,
            cwd=str(cwd) if cwd else None,
            env=env,
            text=True,
            encoding="utf-8",
            errors="replace",
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            check=False,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        output = exc.stdout or ""
        raise RuntimeError(
            f"Command timed out after {timeout}s: {' '.join(command[:3])}\n{str(output)[-1000:]}"
        ) from exc


def node_platform_id() -> str:
    machine = platform.machine().lower()
    if os.name == "nt" and machine in {"amd64", "x86_64"}:
        return "windows-x64"
    if sys.platform == "darwin" and machine == "x86_64":
        return "macos-intel"
    if sys.platform == "darwin" and machine == "arm64":
        return "macos-apple-silicon"
    raise RuntimeError(
        f"AdsPower publishing supports Windows x64, Intel Mac, and Apple Silicon Mac; current platform is {sys.platform}/{machine}."
    )


def node_works(executable: Path) -> bool:
    if not executable.is_file():
        return False
    result = run_quiet([str(executable), "--version"], timeout=30)
    if result.returncode != 0:
        return False
    version = result.stdout.strip().lstrip("v").split(".", 1)[0]
    return version.isdigit() and int(version) >= 18


def cached_node_path(platform_id: str) -> Path:
    asset = NODE_ASSETS[platform_id]
    return CACHE_ROOT / "node" / NODE_VERSION / platform_id / str(asset["executable"])


def adspower_dependencies_ready() -> bool:
    required = [
        ADSPOWER_CLI,
        ADSPOWER_CONFIG_TEMPLATE,
        ADSPOWER_RUNTIME_DIR / "node_modules" / "playwright" / "package.json",
    ]
    return all(path.is_file() for path in required)


def ensure_adspower_payload() -> Path:
    if adspower_dependencies_ready():
        return ADSPOWER_RUNTIME_DIR
    if not ADSPOWER_RUNTIME_ARCHIVE.is_file():
        raise RuntimeError("H is missing its bundled AdsPower runtime archive. Reinstall H from GitHub.")
    if sha256_file(ADSPOWER_RUNTIME_ARCHIVE) != ADSPOWER_RUNTIME_SHA256:
        raise RuntimeError("H bundled AdsPower runtime failed SHA-256 verification. Reinstall H from GitHub.")
    ADSPOWER_RUNTIME_DIR.parent.mkdir(parents=True, exist_ok=True)
    with BootstrapLock():
        if adspower_dependencies_ready():
            return ADSPOWER_RUNTIME_DIR
        staging = Path(
            tempfile.mkdtemp(
                prefix=f".{ADSPOWER_RUNTIME_ID}-",
                dir=str(ADSPOWER_RUNTIME_DIR.parent),
            )
        )
        try:
            with zipfile.ZipFile(ADSPOWER_RUNTIME_ARCHIVE) as archive:
                for member in archive.infolist():
                    assert_archive_member(staging, member.filename)
                archive.extractall(staging)
            required = [
                staging / "src" / "cli.mjs",
                staging / "config.adspower.tiktok.example.json",
                staging / "node_modules" / "playwright" / "package.json",
            ]
            if not all(path.is_file() for path in required):
                raise RuntimeError("H bundled AdsPower runtime archive is incomplete.")
            if ADSPOWER_RUNTIME_DIR.exists():
                shutil.rmtree(ADSPOWER_RUNTIME_DIR)
            os.replace(staging, ADSPOWER_RUNTIME_DIR)
        finally:
            shutil.rmtree(staging, ignore_errors=True)
    if not adspower_dependencies_ready():
        raise RuntimeError("H could not prepare its bundled AdsPower runtime.")
    return ADSPOWER_RUNTIME_DIR


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def assert_archive_member(base: Path, member_name: str, link_name: str = "") -> None:
    target = (base / member_name).resolve()
    try:
        target.relative_to(base.resolve())
    except ValueError as exc:
        raise RuntimeError(f"Unsafe Node archive member: {member_name}") from exc
    if link_name:
        link_target = (target.parent / link_name).resolve()
        try:
            link_target.relative_to(base.resolve())
        except ValueError as exc:
            raise RuntimeError(f"Unsafe Node archive link: {member_name} -> {link_name}") from exc


def extract_node_archive(archive_path: Path, destination: Path) -> Path:
    if archive_path.suffix.lower() == ".zip":
        with zipfile.ZipFile(archive_path) as archive:
            for member in archive.infolist():
                assert_archive_member(destination, member.filename)
            archive.extractall(destination)
    else:
        with tarfile.open(archive_path, "r:gz") as archive:
            for member in archive.getmembers():
                assert_archive_member(destination, member.name, member.linkname if member.issym() or member.islnk() else "")
            archive.extractall(destination)
    roots = [path for path in destination.iterdir() if path.is_dir()]
    if len(roots) != 1:
        raise RuntimeError("The verified Node archive did not contain one runtime directory.")
    return roots[0]


def download_node_runtime(platform_id: str) -> Path:
    asset = NODE_ASSETS[platform_id]
    target_root = CACHE_ROOT / "node" / NODE_VERSION / platform_id
    executable = target_root / str(asset["executable"])
    if node_works(executable):
        return executable

    download_dir = CACHE_ROOT / "downloads"
    download_dir.mkdir(parents=True, exist_ok=True)
    filename = str(asset["filename"])
    archive_path = download_dir / filename
    expected_hash = str(asset["sha256"])
    if not archive_path.is_file() or sha256_file(archive_path) != expected_hash:
        archive_path.unlink(missing_ok=True)
        temporary_path = archive_path.with_suffix(archive_path.suffix + ".part")
        temporary_path.unlink(missing_ok=True)
        url = f"https://nodejs.org/download/release/{NODE_VERSION}/{filename}"
        print(f"H bootstrap: downloading the verified AdsPower runtime for {platform_id}...", flush=True)
        request = urllib.request.Request(url, headers={"User-Agent": "H-Codex-Plugin/0.3"})
        try:
            with urllib.request.urlopen(request, timeout=60) as response, temporary_path.open("wb") as handle:
                shutil.copyfileobj(response, handle, length=1024 * 1024)
        except Exception:
            temporary_path.unlink(missing_ok=True)
            raise
        if sha256_file(temporary_path) != expected_hash:
            temporary_path.unlink(missing_ok=True)
            raise RuntimeError(f"Downloaded Node archive failed SHA-256 verification: {filename}")
        temporary_path.replace(archive_path)

    extract_parent = target_root.parent
    extract_parent.mkdir(parents=True, exist_ok=True)
    staging = Path(tempfile.mkdtemp(prefix=f".{platform_id}-", dir=str(extract_parent)))
    try:
        payload = extract_node_archive(archive_path, staging)
        if target_root.exists():
            shutil.rmtree(target_root)
        shutil.move(str(payload), str(target_root))
    finally:
        shutil.rmtree(staging, ignore_errors=True)
    if not node_works(executable):
        raise RuntimeError(f"The verified Node runtime could not start: {executable}")
    return executable


def ensure_adspower_runtime(*, install: bool) -> dict[str, object]:
    if not adspower_dependencies_ready():
        ensure_adspower_payload()
    if not adspower_dependencies_ready():
        raise RuntimeError("H could not prepare its bundled AdsPower Playwright runtime files.")
    platform_id = node_platform_id()
    candidates: list[tuple[str, Path]] = [("h-cache", cached_node_path(platform_id))]
    system_node = shutil.which("node")
    if system_node:
        candidates.append(("system", Path(system_node)))
    for source, candidate in candidates:
        if node_works(candidate):
            version = run_quiet([str(candidate), "--version"], timeout=30).stdout.strip()
            return {
                "ready": True,
                "platform": platform_id,
                "node": str(candidate),
                "node_version": version,
                "source": source,
                "dependencies": "bundled",
            }
    if not install:
        return {
            "ready": False,
            "platform": platform_id,
            "node": "",
            "source": "missing",
            "dependencies": "bundled",
        }
    with BootstrapLock():
        executable = download_node_runtime(platform_id)
    version = run_quiet([str(executable), "--version"], timeout=30).stdout.strip()
    return {
        "ready": True,
        "platform": platform_id,
        "node": str(executable),
        "node_version": version,
        "source": "downloaded-verified",
        "dependencies": "bundled",
    }


class BootstrapLock:
    def __enter__(self) -> "BootstrapLock":
        CACHE_ROOT.mkdir(parents=True, exist_ok=True)
        deadline = time.monotonic() + 180
        while True:
            try:
                descriptor = os.open(str(LOCK_FILE), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError:
                try:
                    if time.time() - LOCK_FILE.stat().st_mtime > 900:
                        LOCK_FILE.unlink(missing_ok=True)
                        continue
                except OSError:
                    pass
                if time.monotonic() >= deadline:
                    raise RuntimeError("H bootstrap is still locked by another process after 180 seconds.")
                time.sleep(0.25)
                continue
            with os.fdopen(descriptor, "w", encoding="ascii") as handle:
                handle.write(f"{os.getpid()} {int(time.time())}\n")
            return self

    def __exit__(self, _exc_type: object, _exc: object, _traceback: object) -> None:
        try:
            LOCK_FILE.unlink(missing_ok=True)
        except OSError:
            pass


def ensure_python_version() -> None:
    if sys.version_info < MIN_PYTHON:
        current = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        raise RuntimeError(f"H requires Python 3.10 or newer; current Python is {current}.")


def python_works(python: Path) -> bool:
    if not python.exists():
        return False
    result = run_quiet(
        [str(python), "-c", "import sys; raise SystemExit(0 if sys.version_info >= (3, 10) else 1)"],
        timeout=30,
    )
    return result.returncode == 0


def ensure_venv() -> tuple[Path, bool]:
    python = venv_python()
    if python_works(python):
        return python, False
    VENV_DIR.parent.mkdir(parents=True, exist_ok=True)
    print("H bootstrap: creating the reusable Python environment...", flush=True)
    command = [sys.executable, "-m", "venv"]
    if VENV_DIR.exists():
        command.append("--clear")
    command.append(str(VENV_DIR))
    result = run_quiet(command, timeout=180)
    if result.returncode != 0 or not python_works(python):
        raise RuntimeError("H bootstrap failed while creating its Python environment:\n" + result.stdout[-2000:])
    return python, True


def ensure_pip(python: Path) -> None:
    if run_quiet([str(python), "-m", "pip", "--version"], timeout=30).returncode == 0:
        return
    result = run_quiet([str(python), "-m", "ensurepip", "--upgrade"], timeout=180)
    if result.returncode != 0:
        raise RuntimeError("H could not prepare pip in its private environment:\n" + result.stdout[-2000:])


def dependency_marker_current() -> bool:
    try:
        return DEPENDENCY_MARKER.read_text(encoding="ascii").strip() == RUNTIME_ID
    except OSError:
        return False


def missing_imports(python: Path, names: list[str] | None = None) -> list[str]:
    names = names or REQUIRED_IMPORTS
    code = (
        "import importlib.util,json;"
        f"names={names!r};"
        "print(json.dumps([name for name in names if importlib.util.find_spec(name) is None]))"
    )
    result = run_quiet([str(python), "-c", code], timeout=30)
    if result.returncode != 0:
        raise RuntimeError("H could not scan its Python dependencies:\n" + result.stdout[-1000:])
    try:
        value = json.loads(result.stdout.strip().splitlines()[-1])
    except (IndexError, json.JSONDecodeError) as exc:
        raise RuntimeError("H dependency scan returned an unreadable result.") from exc
    return [str(name) for name in value]


def dependencies_ready(python: Path) -> bool:
    return dependency_marker_current() and not missing_imports(python)


def ensure_dependencies(python: Path) -> tuple[bool, list[str], bool]:
    marker_was_current = dependency_marker_current()
    missing_before = missing_imports(python)
    if marker_was_current and not missing_before:
        return False, missing_before, marker_was_current
    ensure_pip(python)
    print("H bootstrap: installing missing runtime dependencies...", flush=True)
    base_command = [
        str(python),
        "-m",
        "pip",
        "install",
        "--disable-pip-version-check",
        "-q",
        "-r",
        str(REQUIREMENTS),
    ]
    result = run_quiet(base_command, timeout=600)
    if result.returncode != 0:
        result = run_quiet([*base_command, "--no-cache-dir"], timeout=600)
    if result.returncode != 0:
        raise RuntimeError("H dependency installation failed:\n" + result.stdout[-2000:])
    missing_after = missing_imports(python)
    if missing_after:
        raise RuntimeError("H dependency installation finished but imports are still missing: " + ", ".join(missing_after))
    DEPENDENCY_MARKER.write_text(RUNTIME_ID, encoding="ascii")
    return True, missing_before, marker_was_current


def bootstrap() -> BootstrapReport:
    packaged_core = packaged_core_path()
    if packaged_core:
        return BootstrapReport(
            python=str(packaged_core),
            environment_created=False,
            dependencies_installed=False,
            missing_before=[],
            marker_was_current=True,
            runtime_source="packaged-executable",
        )
    ensure_python_version()
    with BootstrapLock():
        python, environment_created = ensure_venv()
        dependencies_installed, missing_before, marker_was_current = ensure_dependencies(python)
    return BootstrapReport(
        python=str(python),
        environment_created=environment_created,
        dependencies_installed=dependencies_installed,
        missing_before=missing_before,
        marker_was_current=marker_was_current,
        runtime_source=os.environ.get("H_BOOTSTRAP_PYTHON_SOURCE", "direct-python"),
    )


def desktop_dir() -> Path:
    if os.name == "nt":
        try:
            import winreg

            key_path = r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path) as key:
                value, _ = winreg.QueryValueEx(key, "Desktop")
                if value:
                    return Path(os.path.expandvars(str(value))).expanduser()
        except (OSError, ImportError):
            pass
    return Path.home() / "Desktop"


def desktop_writable() -> bool:
    desktop = desktop_dir()
    try:
        desktop.mkdir(parents=True, exist_ok=True)
        probe = desktop / f".h-write-test-{os.getpid()}"
        probe.write_text("ok", encoding="ascii")
        probe.unlink()
        return True
    except OSError:
        return False


def secret_present(value: str) -> bool:
    cleaned = value.strip().lstrip("\ufeff")
    return bool("".join(character for character in cleaned if character.isprintable() and not character.isspace()))


def clean_secret(value: str) -> str:
    return "".join(
        character
        for character in value.strip().lstrip("\ufeff")
        if character.isprintable() and not character.isspace()
    )


def read_secret_file(path: Path) -> str:
    try:
        return clean_secret(path.read_text(encoding="utf-8-sig"))
    except OSError:
        return ""


def key_file_present(path: Path) -> bool:
    return bool(read_secret_file(path))


def key_sources() -> list[str]:
    sources: list[str] = []
    if secret_present(os.environ.get("H_KIE_API_KEY", "")):
        sources.append("H_KIE_API_KEY")
    if secret_present(os.environ.get("KIE_API_KEY", "")):
        sources.append("KIE_API_KEY")
    if key_file_present(USER_KEY_FILE):
        sources.append("<home>/.codex/secrets/h_kie_api_key.txt")
    if key_file_present(LOCAL_KEY_FILE):
        sources.append("plugin-local .h_api_key")
    return sources


def fastmoss_key_sources() -> list[str]:
    sources: list[str] = []
    if secret_present(os.environ.get("FASTMOSS_API_KEY", "")):
        sources.append("FASTMOSS_API_KEY")
    if secret_present(os.environ.get("H_FASTMOSS_API_KEY", "")):
        sources.append("H_FASTMOSS_API_KEY")
    if key_file_present(FASTMOSS_KEY_FILE):
        sources.append("<home>/.codex/secrets/h_fastmoss_api_key.txt")
    return sources


def fastmoss_api_key() -> tuple[str, str]:
    candidates = [
        ("FASTMOSS_API_KEY", clean_secret(os.environ.get("FASTMOSS_API_KEY", ""))),
        ("H_FASTMOSS_API_KEY", clean_secret(os.environ.get("H_FASTMOSS_API_KEY", ""))),
        ("<home>/.codex/secrets/h_fastmoss_api_key.txt", read_secret_file(FASTMOSS_KEY_FILE)),
    ]
    kie_values = configured_kie_keys()
    for source, value in candidates:
        if value:
            if value in kie_values:
                raise FastMossError(
                    "FastMoss API key is identical to the Kie API key; the credential was likely saved in the wrong key slot.",
                    category="authentication",
                )
            return value, source
    raise FastMossError("FastMoss API key is missing.", category="authentication")


def configured_kie_keys() -> set[str]:
    return {
        value
        for value in (
            clean_secret(os.environ.get("H_KIE_API_KEY", "")),
            clean_secret(os.environ.get("KIE_API_KEY", "")),
            read_secret_file(USER_KEY_FILE),
            read_secret_file(LOCAL_KEY_FILE),
        )
        if value
    }


def write_secret_file(path: Path, value: str) -> None:
    cleaned = clean_secret(value)
    if not cleaned:
        raise RuntimeError("No API key was entered.")
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    descriptor = os.open(str(temporary), os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(cleaned)
        os.chmod(temporary, 0o600)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def plugin_version() -> str:
    manifest = PLUGIN_ROOT / ".codex-plugin" / "plugin.json"
    try:
        return str(json.loads(manifest.read_text(encoding="utf-8"))["version"])
    except (OSError, KeyError, json.JSONDecodeError, TypeError):
        return "unknown"


def write_ready(report: BootstrapReport, checks: dict[str, object]) -> None:
    CACHE_ROOT.mkdir(parents=True, exist_ok=True)
    data = {
        "ready": True,
        "api_verified": True,
        "time": int(time.time()),
        "python": report.python,
        "plugin_root": str(PLUGIN_ROOT),
        "plugin_version": plugin_version(),
        "requirements_hash": RUNTIME_ID,
        "checks": checks,
    }
    READY_FILE.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def ready_cache_valid() -> bool:
    try:
        data = json.loads(READY_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    return bool(
        data.get("ready")
        and data.get("api_verified")
        and data.get("requirements_hash") == RUNTIME_ID
        and data.get("plugin_version") == plugin_version()
    )


def parse_last_json(output: str) -> dict[str, object]:
    lines = output.strip().splitlines()
    for start in range(len(lines)):
        candidate = "\n".join(lines[start:])
        try:
            value = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            return value
    return {}


def local_checks(report: BootstrapReport, *, offline: bool, forwarded_args: list[str]) -> dict[str, object]:
    sources = key_sources()
    if "--api-key" in forwarded_args:
        sources = ["--api-key", *sources]
    packaged = report.runtime_source == "packaged-executable"
    runtime_path = Path(report.python)
    return {
        "python_version": ".".join(str(value) for value in sys.version_info[:3]),
        "python_supported": sys.version_info >= MIN_PYTHON,
        "requirements": REQUIREMENTS.exists(),
        "dependencies": runtime_path.is_file() if packaged else dependencies_ready(runtime_path),
        "main_script": runtime_path.is_file() if packaged else MAIN_SCRIPT.exists(),
        "desktop": str(desktop_dir()),
        "desktop_writable": desktop_writable(),
        "kie_key_sources": sources,
        "fastmoss_key_sources": fastmoss_key_sources(),
        "runtime": report.python if packaged else str(VENV_DIR),
        "portable": packaged,
        "offline": offline,
    }


def core_command(report: BootstrapReport, arguments: list[str]) -> list[str]:
    if report.runtime_source == "packaged-executable":
        return [report.python, *arguments]
    return [report.python, str(MAIN_SCRIPT), *arguments]


def run_api_doctor(report: BootstrapReport, forwarded_args: list[str]) -> tuple[int, dict[str, object], str]:
    result = run_quiet(
        core_command(report, ["doctor", *forwarded_args]),
        cwd=PLUGIN_ROOT,
        timeout=60,
    )
    return result.returncode, parse_last_json(result.stdout), result.stdout


def doctor(*, offline: bool = False, forwarded_args: list[str] | None = None) -> int:
    report = bootstrap()
    forwarded_args = forwarded_args or []
    checks = local_checks(report, offline=offline, forwarded_args=forwarded_args)
    try:
        checks["adspower_runtime"] = ensure_adspower_runtime(install=not offline)
    except Exception as exc:
        if not offline:
            raise
        checks["adspower_runtime"] = {"ready": False, "error": safe_reason(exc)}
    sources = checks["kie_key_sources"]
    local_ready = all(
        bool(checks[name])
        for name in ("python_supported", "requirements", "dependencies", "main_script", "desktop_writable")
    ) and (bool(sources) or offline)
    api_ready = offline
    if local_ready and not offline:
        code, api_result, raw_output = run_api_doctor(report, forwarded_args)
        checks["kie_api"] = api_result or {"ready": False, "message": raw_output[-1000:]}
        api_ready = code == 0 and bool(api_result.get("ready"))
    ready = local_ready and api_ready
    if ready and not offline:
        write_ready(report, checks)
    print_json({"ready": ready, "bootstrap": asdict(report), "checks": checks})
    return 0 if ready else 1


def model_catalog(report: BootstrapReport) -> dict[str, Any]:
    result = run_quiet(core_command(report, ["catalog"]), cwd=PLUGIN_ROOT, timeout=60)
    catalog = parse_last_json(result.stdout)
    if result.returncode != 0 or not catalog:
        raise RuntimeError("H could not read its model catalog:\n" + result.stdout[-1000:])
    return catalog


def catalog_lines(items: list[dict[str, Any]], kind: str) -> str:
    lines: list[str] = []
    for item in items:
        choice = item.get("choice", "")
        name = item.get("name", item.get("model", ""))
        if kind == "image":
            limit = item.get("max_reference_images", 0)
            detail = f"0 张参考图=文生图，1-{limit} 张=多图参考"
        elif kind == "video":
            fixed = item.get("fixed_seconds")
            maximum = item.get("max_seconds")
            duration = f"固定约 {fixed} 秒" if fixed else f"最长 {maximum} 秒" if maximum else "时长按 Kie 当前支持"
            detail = f"{duration}；{item.get('input_rule', '')}"
        else:
            detail = ""
        suffix = f"（{detail}）" if detail else ""
        lines.append(f"{choice}. {name}{suffix}")
    return "\n".join(lines)


def protocol_display(state: str, catalog: dict[str, Any]) -> str:
    state = state.strip().lower().replace("_", "-")
    text_models = catalog_lines(list(catalog.get("text", [])), "text")
    image_models = catalog_lines(list(catalog.get("image", [])), "image")
    video_items = list(catalog.get("video", []))
    batch_video_items = [item for item in video_items if not str(item.get("input_rule", "")).startswith("已有Kie")]
    video_models = catalog_lines(video_items, "video")
    batch_video_models = catalog_lines(batch_video_items, "video")
    menus = {
        "mode": MODE_MENU,
        "pid": (
            "请发送一个或多个商品 PID。H 会从 FastMoss 拉取商品主图和标题，"
            "再让 AI 同时分析图片与标题后生成视频。多个 PID 可用空格、逗号或换行分隔。"
        ),
        "generate-mode": GENERATE_MENU,
        "batch-root": "请发送要批处理的根文件夹路径。H 会递归扫描整个根目录，并把全部合格图片放进同一个并发池。",
        "batch-image": (
            "请一次回复：图片模型编号 / 分辨率编号 / 比例编号 / 图片反推文本模型编号 / 图片反推元提示词。\n\n"
            f"图片模型：\n{image_models}\n\n"
            "分辨率：\n1. 1K\n2. 2K\n3. 4K\n\n"
            "比例：\n1. 9:16\n2. 16:9\n\n"
            f"图片反推文本模型：\n{text_models}\n\n"
            "图片反推元提示词直接回车使用中文默认值。"
        ),
        "batch-video": (
            "请一次回复：视频模型编号 / 时长秒数 / 分辨率 / 比例编号 / 视频反推文本模型编号 / 视频反推元提示词。\n\n"
            f"视频模型：\n{batch_video_models}\n\n"
            "分辨率：480p / 720p / 1080p（Veo 仅 720p/1080p；Gemini Omni 由模型决定）\n"
            "比例：1. 9:16  2. 16:9\n\n"
            f"视频反推文本模型：\n{text_models}\n\n"
            "视频反推元提示词直接回车使用中文默认值。"
        ),
        "pid-video": (
            "FastMoss 商品主图和标题已准备好。请一次回复：视频模型编号 / 时长秒数 / 分辨率 / "
            "比例编号 / AI 分析文本模型编号 / 视频元提示词。\n\n"
            f"视频模型：\n{batch_video_models}\n\n"
            "分辨率：480p / 720p / 1080p（Veo 仅 720p/1080p；Gemini Omni 由模型决定）\n"
            "比例：1. 9:16  2. 16:9\n\n"
            f"AI 分析文本模型：\n{text_models}\n\n"
            "AI 会把原始商品图和 FastMoss 标题一起分析；元提示词直接回车使用中文默认值。"
        ),
        "single-kind": "请选择单处理类型，回复编号即可：\n1. 文本\n2. 图像\n3. 视频",
        "single-text": f"请选择文本模型并发送 prompt：\n{text_models}",
        "single-image": (
            "请一次发送：图片模型编号 / 分辨率编号 / 比例编号 / prompt / 参考图片。参考图片可不传，也可一次传多张。\n\n"
            f"图片模型：\n{image_models}\n\n"
            "分辨率：\n1. 1K\n2. 2K\n3. 4K\n\n"
            "比例：\n1. 9:16\n2. 16:9"
        ),
        "single-video": (
            "请一次发送：视频模型编号 / 时长秒数 / 分辨率 / 比例编号 / prompt / 所需图片、视频、音频或 Kie 任务 ID。\n\n"
            f"视频模型：\n{video_models}"
        ),
        "post-images": "请选择下一步：\n1. 继续生成视频\n2. 只重试失败项\n3. 处理新的文件夹\n4. 结束",
        "post-videos": "请选择下一步：\n1. 发布本次生成的视频\n2. 只重试失败项\n3. 处理新的文件夹\n4. 结束",
        "post-single": (
            "请选择下一步：\n"
            "1. 重试或继续当前任务（已提交任务只查询，不重复提交）\n"
            "2. 继续新的单处理\n3. 切换到批处理\n4. 结束"
        ),
        "post-single-video": (
            "请选择下一步：\n"
            "1. 发布本次生成的视频\n"
            "2. 重试或继续当前任务（已提交任务只查询，不重复提交）\n"
            "3. 继续新的单处理\n4. 切换到批处理\n5. 结束"
        ),
        "publish-source": (
            "请选择发布来源，回复编号即可：\n"
            "1. H 已生成的视频结果\n"
            "2. 普通视频文件夹\n"
            "3. 已有 XLSX/CSV 发布计划表"
        ),
        "publish-plan": (
            "请一次发送：视频或 H 结果目录 / AdsPower 环境编号（可多个） / 首次发布时间 / 半小时间隔编号 / "
            "文案模板 / 标签 / 是否按数字 PID 挂商品 / 时区。\n"
            "间隔：1. 30 分钟  2. 60 分钟  3. 90 分钟  4. 120 分钟，也可填写其他 30 分钟倍数。\n"
            "文案模板可使用 {pid}、{index}、{filename}；PID 视频必须挂同一个完整数字 PID，绝不按标题匹配。"
        ),
        "publish-file": "请发送现有 XLSX 或 CSV 发布计划表的完整路径。",
        "publish-review": "发布计划已生成并校验。请选择下一步：\n1. 预览上传（不点击最终发布）\n2. 修改计划\n3. 结束",
        "publish-confirm": (
            "预览已完成。确认账号、文案、商品 PID 和时间均正确后，输入 FABU 正式发布；\n"
            "回复 2 修改计划，回复 3 结束。"
        ),
        "post-publish": "请选择下一步：\n1. 查看发布结果与日志\n2. 新的发布任务\n3. 返回生成\n4. 结束",
    }
    if state not in menus:
        raise ValueError(f"Unknown H protocol state: {state}")
    return menus[state]


def protocol(state: str) -> int:
    report = bootstrap()
    catalog: dict[str, Any] = {}
    normalized = state.strip().lower().replace("_", "-")
    if normalized not in {
        "mode",
        "pid",
        "generate-mode",
        "batch-root",
        "single-kind",
        "post-images",
        "post-videos",
        "post-single",
        "post-single-video",
        "publish-source",
        "publish-plan",
        "publish-file",
        "publish-review",
        "publish-confirm",
        "post-publish",
    }:
        catalog = model_catalog(report)
    print_json(
        {
            "ready": True,
            "protocol_version": PROTOCOL_VERSION,
            "state": normalized,
            "bootstrap": asdict(report),
            "display_text": protocol_display(normalized, catalog),
        }
    )
    return 0


def setup_status(report: BootstrapReport) -> str:
    if report.runtime_source == "packaged-executable":
        return "内置运行环境已就绪，无需安装 Python。"
    if report.environment_created or report.dependencies_installed:
        return "首次环境准备完成，缺失依赖已自动安装。"
    return "环境检查通过，H 已准备好。"


def category_label(category: str) -> str:
    labels = {
        "authentication": "Kie 密钥无效或已失效",
        "quota": "Kie 额度不足",
        "validation": "请求参数不符合模型要求",
        "rate_limit": "Kie 接口限流",
        "maintenance": "Kie 服务维护中",
        "provider": "上游模型服务异常",
        "network": "网络、代理或 TLS 连接异常",
        "runtime": "本地运行环境异常",
    }
    return labels.get(category, "未分类错误")


def safe_reason(value: object) -> str:
    text = " ".join(str(value or "未知原因").split())
    return text[:500]


def start(
    *,
    offline: bool = False,
    force_check: bool = False,
    capability: str = "menu",
    forwarded_args: list[str] | None = None,
) -> int:
    capability = capability.strip().lower()
    if capability not in {"menu", "pid", "generate", "publish"}:
        raise ValueError(f"Unsupported H capability: {capability}")
    report = bootstrap()
    forwarded_args = forwarded_args or []
    checks = local_checks(report, offline=offline, forwarded_args=forwarded_args)
    if capability == "publish":
        try:
            checks["adspower_runtime"] = ensure_adspower_runtime(install=not offline)
        except Exception as exc:
            if not offline:
                raise
            checks["adspower_runtime"] = {"ready": False, "error": safe_reason(exc)}
    else:
        checks["adspower_runtime"] = {"ready": None, "skipped": True}
    status = setup_status(report)
    if not checks["desktop_writable"]:
        display = f"{GREETING}\n\nH 自动环境准备失败。\n归因：桌面目录不可写。\n请修复桌面目录权限后重新调用 H。"
        print_json({"ready": False, "state": "setup-error", "bootstrap": asdict(report), "checks": checks, "display_text": display})
        return 0
    if offline:
        print_json(
            {
                "ready": True,
                "state": "mode",
                "protocol_version": PROTOCOL_VERSION,
                "bootstrap": asdict(report),
                "checks": checks,
                "display_text": f"{GREETING}\n\n{status}\n\n{MODE_MENU}",
            }
        )
        return 0
    if capability == "menu":
        print_json(
            {
                "ready": True,
                "state": "mode",
                "protocol_version": PROTOCOL_VERSION,
                "bootstrap": asdict(report),
                "checks": checks,
                "display_text": f"{GREETING}\n\n{status}\n\n{MODE_MENU}",
            }
        )
        return 0
    if capability == "pid":
        if not checks["fastmoss_key_sources"]:
            display = (
                f"{GREETING}\n\n{status}\n\n"
                "尚未检测到 FastMoss API Key。请设置 FASTMOSS_API_KEY，或运行一次 H 的 "
                "set-fastmoss-key；密钥只保存在当前用户的 .codex/secrets 目录。"
            )
            print_json(
                {
                    "ready": False,
                    "state": "fastmoss-key-required",
                    "bootstrap": asdict(report),
                    "checks": checks,
                    "display_text": display,
                }
            )
            return 0
        print_json(
            {
                "ready": True,
                "state": "pid",
                "protocol_version": PROTOCOL_VERSION,
                "bootstrap": asdict(report),
                "checks": checks,
                "display_text": f"{GREETING}\n\n{status}\n\n{protocol_display('pid', {})}",
            }
        )
        return 0
    if capability == "publish":
        runtime = checks.get("adspower_runtime")
        if not isinstance(runtime, dict) or not runtime.get("ready"):
            raise RuntimeError("AdsPower publishing runtime is not ready.")
        print_json(
            {
                "ready": True,
                "state": "publish-source",
                "protocol_version": PROTOCOL_VERSION,
                "bootstrap": asdict(report),
                "checks": checks,
                "display_text": f"{GREETING}\n\n{status}\n\n{protocol_display('publish-source', {})}",
            }
        )
        return 0
    if not checks["kie_key_sources"]:
        display = (
            f"{GREETING}\n\n{status}\n\n"
            "尚未检测到 Kie API Key。请先设置一次密钥；H 不会把密钥写入仓库或输出日志。"
        )
        print_json({"ready": False, "state": "key-required", "bootstrap": asdict(report), "checks": checks, "display_text": display})
        return 0
    if ready_cache_valid() and not force_check:
        print_json(
            {
                "ready": True,
                "state": "generate-mode",
                "protocol_version": PROTOCOL_VERSION,
                "bootstrap": asdict(report),
                "checks": checks,
                "display_text": f"{GREETING}\n\n{status}\n\n{GENERATE_MENU}",
            }
        )
        return 0
    code, api_result, raw_output = run_api_doctor(report, forwarded_args)
    checks["kie_api"] = api_result or {"ready": False, "message": raw_output[-1000:]}
    if code == 0 and api_result.get("ready"):
        write_ready(report, checks)
        print_json(
            {
                "ready": True,
                "state": "generate-mode",
                "protocol_version": PROTOCOL_VERSION,
                "bootstrap": asdict(report),
                "checks": checks,
                "display_text": f"{GREETING}\n\n{status}\n\n{GENERATE_MENU}",
            }
        )
        return 0
    category = str(api_result.get("error_category") or api_result.get("category") or "runtime")
    reason = safe_reason(api_result.get("error") or api_result.get("message") or raw_output[-500:])
    display = (
        f"{GREETING}\n\n本地环境已经准备完成，但 Kie 验证失败。\n"
        f"归因：{category_label(category)}。\n原因：{reason}\n"
        "请修正后重新调用 H；验证失败时不会提交任何生成任务。"
    )
    print_json(
        {
            "ready": False,
            "state": "setup-error",
            "error_category": category,
            "bootstrap": asdict(report),
            "checks": checks,
            "display_text": display,
        }
    )
    return 0


def default_fastmoss_work_dir() -> Path:
    return desktop_dir() / "H返回结果_PID"


def fastmoss_product_display(saved: dict[str, Any]) -> str:
    lines = [
        f"FastMoss PID 查询完成：成功 {saved['success']}，可生成视频 {saved['ready_for_generation']}，未找到 {saved['not_found']}。"
    ]
    for item in list(saved["results"])[:10]:
        pid = str(item.get("pid") or "")
        product = item.get("product") if isinstance(item.get("product"), dict) else {}
        if item.get("state") == "not_found":
            lines.append(f"- {pid}：未找到商品")
            continue
        title = " ".join(str(product.get("title") or "未返回标题").split())[:120]
        image_status = "主图已保存" if item.get("reference_image") else "主图下载失败，请上传图片"
        lines.append(f"- {pid}：{title}；{image_status}")
    if len(saved["results"]) > 10:
        lines.append(f"其余 {len(saved['results']) - 10} 个结果已写入结果目录。")
    lines.append(f"结果目录：{saved['generation_root']}")
    if saved["ready_for_generation"]:
        lines.append("下一步：让 AI 同时分析已保存的商品主图和 FastMoss 标题，然后生成视频。")
    return "\n".join(lines)


def fastmoss_error_display(exc: Exception) -> str:
    category = exc.category if isinstance(exc, FastMossError) else "runtime"
    labels = {
        "authentication": "FastMoss 密钥无效、已失效或未配置",
        "quota": "FastMoss 调用次数或额度不足",
        "permission": "当前 FastMoss Key 没有商品查询权限",
        "rate_limit": "FastMoss 接口正在限流",
        "validation": "PID、上传图片或请求参数不符合要求",
        "not_found": "FastMoss 没有找到对应商品",
        "invalid_result": "FastMoss 返回的商品主图不可用",
        "network": "网络、代理或 TLS 连接异常",
        "provider": "FastMoss 服务暂时异常",
        "runtime": "本地运行环境异常",
    }
    lines = ["FastMoss PID 查询失败。", f"归因：{labels.get(category, '未分类错误')}。", f"原因：{safe_reason(exc)}"]
    if isinstance(exc, FastMossError) and exc.request_id:
        lines.append(f"请求编号：{exc.request_id}")
    lines.append("本次没有提交任何 Kie 生成或 TikTok 发布任务。")
    return "\n".join(lines)


def build_fastmoss_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="h_run fastmoss", description="H FastMoss PID workflow")
    commands = parser.add_subparsers(dest="fastmoss_command", required=True)
    product = commands.add_parser("product")
    product.add_argument("--pid", action="append", required=True)
    product.add_argument(
        "--media",
        action="append",
        default=[],
        help="Optional uploaded product image. Repeat once per PID in the same order.",
    )
    product.add_argument("--work-dir", default=str(default_fastmoss_work_dir()))
    product.add_argument("--timeout", type=int, default=60)
    product.add_argument("--skip-image-download", action="store_true")
    commands.add_parser("status")
    return parser


def map_pid_reference_images(pids: list[str], media_values: list[str]) -> dict[str, Path]:
    if not media_values:
        return {}
    if len(media_values) != len(pids):
        raise FastMossError(
            "Uploaded image count must exactly match PID count so every product keeps a one-to-one mapping.",
            category="validation",
        )
    mapped: dict[str, Path] = {}
    for pid, value in zip(pids, media_values, strict=True):
        source, _extension = validate_local_reference_image(Path(value))
        mapped[pid] = source
    return mapped


def fastmoss_command(argv: list[str]) -> int:
    args = build_fastmoss_parser().parse_args(argv)
    if args.fastmoss_command == "status":
        sources = fastmoss_key_sources()
        print_json(
            {
                "ready": bool(sources),
                "stage": "fastmoss-status",
                "key_sources": sources,
                "key_file": str(FASTMOSS_KEY_FILE),
            }
        )
        return 0 if sources else 1
    pids = normalize_pids(args.pid)
    reference_images = map_pid_reference_images(pids, args.media)
    api_key, key_source = fastmoss_api_key()
    query = query_products(pids, api_key, timeout=args.timeout)
    saved = save_product_results(
        query,
        Path(args.work_dir),
        download_images=not args.skip_image_download,
        reference_images=reference_images,
    )
    ready_for_generation = int(saved["ready_for_generation"])
    payload = {
        "ready": ready_for_generation > 0,
        "stage": "fastmoss-product",
        "source": "FastMoss",
        "key_source": key_source,
        **saved,
        "next_state": "pid-video" if ready_for_generation > 0 else "pid",
        "display_text": fastmoss_product_display(saved),
    }
    print_json(payload)
    return 0 if ready_for_generation > 0 else 2


def default_adspower_work_dir() -> Path:
    return desktop_dir() / "H返回结果_发布"


def initialize_adspower_workspace(work_dir: Path) -> dict[str, str]:
    work_dir = work_dir.expanduser().resolve()
    log_dir = work_dir / "logs"
    artifacts_dir = work_dir / "artifacts"
    for directory in (work_dir, log_dir, artifacts_dir):
        directory.mkdir(parents=True, exist_ok=True)
    config_path = work_dir / "config.json"
    schedule_path = work_dir / "schedule.xlsx"
    csv_path = work_dir / "schedule.csv"
    if not config_path.exists():
        shutil.copy2(ADSPOWER_CONFIG_TEMPLATE, config_path)
    if not schedule_path.exists() and ADSPOWER_SCHEDULE_TEMPLATE.is_file():
        shutil.copy2(ADSPOWER_SCHEDULE_TEMPLATE, schedule_path)
    if not csv_path.exists() and ADSPOWER_SCHEDULE_CSV.is_file():
        shutil.copy2(ADSPOWER_SCHEDULE_CSV, csv_path)
    return {
        "work_dir": str(work_dir),
        "config": str(config_path),
        "schedule": str(schedule_path if schedule_path.exists() else csv_path),
        "logs": str(log_dir),
        "artifacts": str(artifacts_dir),
    }


def video_file_valid(path: Path) -> bool:
    try:
        if not path.is_file() or path.stat().st_size < 16:
            return False
        with path.open("rb") as handle:
            header = handle.read(16)
    except OSError:
        return False
    return header[4:8] == b"ftyp" or header.startswith(b"\x1aE\xdf\xa3")


def iter_result_records(value: object) -> list[dict[str, object]]:
    if isinstance(value, dict):
        return [value]
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def discover_publish_videos(video_root: Path) -> list[dict[str, str]]:
    root = video_root.expanduser().resolve()
    if not root.exists():
        raise FileNotFoundError(f"Video or H result directory does not exist: {root}")
    if root.is_file():
        if root.suffix.lower() not in VIDEO_EXTENSIONS or not video_file_valid(root):
            raise ValueError(f"The selected file is not a valid MP4/MOV/WebM video: {root}")
        return [{"path": str(root), "pid": root.stem, "filename": root.name}]

    discovered: dict[str, dict[str, str]] = {}
    for json_path in root.rglob("*.json"):
        try:
            value = json.loads(json_path.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError, UnicodeError):
            continue
        for record in iter_result_records(value):
            if str(record.get("state", "")).lower() != "success":
                continue
            candidate_value = record.get("video_path")
            if not candidate_value and str(record.get("kind", "")).lower() == "video":
                candidate_value = record.get("output_path")
            if not candidate_value:
                continue
            candidate = Path(str(candidate_value)).expanduser()
            if not candidate.is_absolute():
                candidate = (json_path.parent / candidate).resolve()
            else:
                candidate = candidate.resolve()
            if candidate.suffix.lower() not in VIDEO_EXTENSIONS or not video_file_valid(candidate):
                continue
            key = os.path.normcase(str(candidate))
            record_pid = str(record.get("pid") or candidate.stem).strip()
            existing = discovered.get(key)
            if existing and existing["pid"] != record_pid:
                raise ValueError(
                    f"Conflicting PID records for {candidate}: {existing['pid']} and {record_pid}"
                )
            discovered[key] = {
                "path": str(candidate),
                "pid": record_pid,
                "filename": candidate.name,
            }
    for candidate in root.rglob("*"):
        if candidate.suffix.lower() not in VIDEO_EXTENSIONS or not video_file_valid(candidate):
            continue
        resolved = candidate.resolve()
        key = os.path.normcase(str(resolved))
        discovered.setdefault(
            key,
            {"path": str(resolved), "pid": resolved.stem, "filename": resolved.name},
        )
    return sorted(discovered.values(), key=lambda item: (item["pid"].lower(), item["path"].lower()))


def parse_profiles(values: list[str]) -> list[str]:
    profiles: list[str] = []
    for value in values:
        for item in value.replace("，", ",").split(","):
            cleaned = item.strip()
            if cleaned and cleaned not in profiles:
                profiles.append(cleaned)
    if not profiles:
        raise ValueError("At least one AdsPower profile number is required.")
    return profiles


def render_caption(template: str, item: dict[str, str], index: int) -> str:
    return (
        template.replace("{pid}", item["pid"])
        .replace("{index}", str(index))
        .replace("{filename}", item["filename"])
    )


def create_adspower_plan(args: argparse.Namespace, work_dir: Path) -> dict[str, object]:
    videos = discover_publish_videos(Path(args.video_root))
    if not videos:
        raise ValueError("No valid generated video files were found. PNG/JPEG files renamed to .mp4 are rejected.")
    profiles = parse_profiles(args.profile_no)
    try:
        start_at = datetime.strptime(args.start_at.strip(), "%Y-%m-%d %H:%M")
    except ValueError as exc:
        raise ValueError("Start time must use YYYY-MM-DD HH:MM.") from exc
    if args.interval_minutes < 30 or args.interval_minutes % 30:
        raise ValueError("Schedule interval must be a positive multiple of 30 minutes.")
    invalid_mappings = [
        f"{item['filename']} -> {item['pid']}"
        for item in videos
        if not item["pid"].isdigit() or Path(item["path"]).stem != item["pid"]
    ]
    if args.attach_pid and invalid_mappings:
        raise ValueError(
            "Every video filename and result record must carry the same exact numeric PID before product attachment. "
            "Invalid mappings: " + ", ".join(invalid_mappings[:20])
        )
    plan_path = (work_dir / args.plan_name).resolve()
    plan_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "启用",
        "环境编号",
        "视频路径",
        "文案",
        "标签",
        "商品PID",
        "预定时间",
        "时区",
        "发布模式",
        "任务ID",
    ]
    mappings: list[dict[str, str]] = []
    with plan_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for offset, item in enumerate(videos):
            index = offset + 1
            pid = item["pid"]
            product_pid = pid if args.attach_pid and pid.isdigit() else ""
            profile = profiles[offset % len(profiles)]
            scheduled_at = (start_at + timedelta(minutes=offset * args.interval_minutes)).strftime("%Y-%m-%d %H:%M")
            writer.writerow(
                {
                    "启用": "yes",
                    "环境编号": profile,
                    "视频路径": item["path"],
                    "文案": render_caption(args.caption_template, item, index),
                    "标签": args.hashtags.strip(),
                    "商品PID": product_pid,
                    "预定时间": scheduled_at,
                    "时区": args.timezone.strip(),
                    "发布模式": args.publish_mode,
                    "任务ID": f"h-{index:04d}-{''.join(character for character in pid if character.isalnum())[:24] or 'video'}",
                }
            )
            mappings.append(
                {
                    "video": item["path"],
                    "pid": product_pid,
                    "profile": profile,
                    "scheduled_at": scheduled_at,
                }
            )
    return {
        "ready": True,
        "stage": "publish-plan",
        "plan": str(plan_path),
        "video_root": str(Path(args.video_root).expanduser().resolve()),
        "videos": len(videos),
        "profiles": profiles,
        "start_at": start_at.strftime("%Y-%m-%d %H:%M"),
        "interval_minutes": args.interval_minutes,
        "attach_pid": bool(args.attach_pid),
        "mappings": mappings,
        "next_state": "publish-review",
    }


def adspower_error_category(value: object) -> str:
    text = str(value or "").lower()
    if "captcha" in text or "verification" in text or "验证码" in text or "安全验证" in text:
        return "manual_verification"
    if "login" in text or "sign in" in text or "登录" in text:
        return "login_required"
    if "adspower api" in text or "127.0.0.1:50325" in text:
        return "adspower_not_running"
    if "计划表" in text or "publish-plan" in text or "validation" in text or "does not exist" in text:
        return "validation"
    if "timed out" in text or "timeout" in text:
        return "timeout"
    if "final click occurred" in text:
        return "publish_unverified"
    return "runtime"


def load_json_value(path: Path) -> object:
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError, UnicodeError):
        return None


def run_adspower_node(
    node: Path,
    arguments: list[str],
    *,
    work_dir: Path,
    timeout: int = 21600,
    env_overrides: dict[str, str] | None = None,
) -> subprocess.CompletedProcess[str]:
    return run_quiet(
        [str(node), *arguments],
        cwd=work_dir,
        timeout=timeout,
        env_overrides=env_overrides,
    )


def schedule_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def schedule_datetime(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    text = schedule_text(value)
    for pattern in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(text[:16], pattern).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return ""


def read_schedule_rows(input_path: Path) -> list[list[object]]:
    extension = input_path.suffix.lower()
    if extension == ".csv":
        try:
            with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
                return [list(row) for row in csv.reader(handle)]
        except UnicodeDecodeError as exc:
            raise ValueError("CSV publish plans must use UTF-8 encoding.") from exc
    if extension != ".xlsx":
        raise ValueError("Only .xlsx and UTF-8 .csv publish plans are supported.")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("H did not finish installing its safe XLSX parser. Run H start again.") from exc
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = workbook["发布计划"] if "发布计划" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def prepare_adspower_tasks(input_path: Path, output_path: Path, mode: str) -> dict[str, object]:
    rows = read_schedule_rows(input_path)
    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if any(schedule_text(value) in {"启用", "enabled"} for value in row)
        ),
        -1,
    )
    if header_index < 0:
        raise ValueError("Cannot find the publish-plan header row.")
    headers = [schedule_text(value) for value in rows[header_index]]

    def pick(row: dict[str, object], *names: str) -> object:
        for name in names:
            if name in row and row[name] is not None:
                return row[name]
        return ""

    tasks: list[dict[str, object]] = []
    errors: list[str] = []
    for index in range(header_index + 1, len(rows)):
        values = rows[index]
        row = {header: values[column] if column < len(values) else "" for column, header in enumerate(headers)}
        enabled = schedule_text(pick(row, "启用", "enabled")).lower()
        if not enabled or enabled in {"no", "n", "0", "false", "否"}:
            continue
        line = index + 1
        profile_no = schedule_text(pick(row, "环境编号", "profileNo"))
        raw_video = schedule_text(pick(row, "视频路径", "videoPath")).strip("\"'")
        video_path = Path(raw_video).expanduser() if raw_video else Path()
        if raw_video and not video_path.is_absolute():
            video_path = input_path.parent / video_path
        video_path = video_path.resolve() if raw_video else video_path
        scheduled_at = schedule_datetime(pick(row, "预定时间", "scheduledAt"))
        timezone = schedule_text(pick(row, "时区", "timezone", "timeZone"))
        caption = schedule_text(pick(row, "文案", "caption"))
        hashtags = schedule_text(pick(row, "标签", "hashtags"))
        raw_pid = pick(row, "商品PID", "productPid", "商品关键词", "productKeyword")
        if raw_pid not in {None, ""} and input_path.suffix.lower() == ".xlsx" and not isinstance(raw_pid, str):
            product_pid = ""
            errors.append(f"第 {line} 行商品PID被 Excel 当成数字，必须把该单元格设为文本后重新填写完整 PID")
        else:
            product_pid = schedule_text(raw_pid).lstrip("'")
        publish_mode = schedule_text(pick(row, "发布模式", "publishMode")).lower()
        task_id = schedule_text(pick(row, "任务ID", "taskId")) or f"profile-{profile_no}-row-{line}"

        if not profile_no:
            errors.append(f"第 {line} 行缺少环境编号")
        if not raw_video:
            errors.append(f"第 {line} 行缺少视频路径")
        elif not video_file_valid(video_path):
            errors.append(f"第 {line} 行视频不存在或文件头不是有效 MP4/MOV/WebM：{video_path}")
        if not scheduled_at:
            errors.append(f"第 {line} 行预定时间无效")
        if product_pid and not product_pid.isdigit():
            errors.append(f"第 {line} 行商品PID必须是完整数字，不能使用标题、模糊关键词或科学计数法")
        tasks.append(
            {
                "id": task_id,
                "profileNo": profile_no,
                "videoPath": str(video_path),
                "description": " ".join(value for value in (caption, hashtags) if value),
                "productPid": product_pid,
                "productKeyword": product_pid,
                "scheduledAt": scheduled_at,
                "timezone": timezone,
                "publish": mode == "publish" and publish_mode != "draft",
            }
        )
    if errors:
        raise ValueError("计划表校验失败：\n- " + "\n- ".join(errors))
    if not tasks:
        raise ValueError("计划表中没有启用的任务。")
    output_path.write_text(json.dumps(tasks, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {"input": str(input_path), "tasks": len(tasks), "mode": mode}


def resolve_adspower_input(args: argparse.Namespace, work_dir: Path, workspace: dict[str, str]) -> Path:
    input_path = Path(args.input_file).expanduser().resolve() if args.input_file else work_dir / "schedule.generated.csv"
    if not input_path.is_file():
        fallback = Path(workspace["schedule"])
        input_path = fallback if not args.input_file and fallback.is_file() else input_path
    if not input_path.is_file():
        raise FileNotFoundError(f"Publish plan does not exist: {input_path}")
    return input_path


def execute_adspower_plan(
    command: str,
    args: argparse.Namespace,
    node: Path,
    work_dir: Path,
    workspace: dict[str, str],
) -> int:
    if command == "publish" and args.publish_code != "FABU":
        raise ValueError("Formal publishing requires the exact confirmation code FABU.")
    input_path = resolve_adspower_input(args, work_dir, workspace)

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    tasks_path = work_dir / "tasks" / f"{command}-{stamp}.json"
    tasks_path.parent.mkdir(parents=True, exist_ok=True)
    prepare_adspower_tasks(input_path, tasks_path, command)

    report_path = work_dir / "logs" / f"{command}-{stamp}.json"
    log_path = work_dir / "logs" / f"{command}-{stamp}.log"
    result = run_adspower_node(
        node,
        [
            str(ADSPOWER_CLI),
            "--config",
            workspace["config"],
            "--tasks",
            str(tasks_path),
            "--report",
            str(report_path),
        ],
        work_dir=work_dir,
        env_overrides={"ADSPOWER_PUBLISH_HEADLESS": "0" if args.visible else "1"},
    )
    log_path.write_text(result.stdout, encoding="utf-8")
    report = load_json_value(report_path)
    entries = report if isinstance(report, list) else []
    failed = sum(1 for item in entries if isinstance(item, dict) and item.get("status") not in {"preview_ready", "published"})
    payload = {
        "ready": result.returncode == 0 and failed == 0,
        "stage": command,
        "work_dir": str(work_dir),
        "input_file": str(input_path),
        "tasks": str(tasks_path),
        "report": str(report_path),
        "log": str(log_path),
        "total": len(entries),
        "success": len(entries) - failed,
        "failed": failed,
        "results": entries,
        "next_state": "publish-confirm" if command == "preview" and failed == 0 else "post-publish",
    }
    if result.returncode != 0 and not entries:
        payload["error_category"] = adspower_error_category(result.stdout)
        payload["error"] = safe_reason(result.stdout[-1500:])
    print_json(payload)
    return result.returncode


def add_adspower_work_dir(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--work-dir", default=str(default_adspower_work_dir()))


def build_adspower_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="h_run adspower", description="H AdsPower TikTok publishing workflow")
    commands = parser.add_subparsers(dest="ads_command", required=True)
    add_adspower_work_dir(commands.add_parser("init"))

    profiles = commands.add_parser("profiles")
    add_adspower_work_dir(profiles)

    check = commands.add_parser("check")
    add_adspower_work_dir(check)
    check.add_argument("--profile-no", action="append", default=[])
    check.add_argument("--concurrency", type=int, default=3)
    check.add_argument("--visible", action="store_true")

    plan = commands.add_parser("plan")
    add_adspower_work_dir(plan)
    plan.add_argument("--video-root", required=True)
    plan.add_argument("--profile-no", action="append", required=True)
    plan.add_argument("--start-at", required=True)
    plan.add_argument("--interval-minutes", type=int, default=60)
    plan.add_argument("--caption-template", default="{pid}")
    plan.add_argument("--hashtags", nargs="?", const="", default="")
    plan.add_argument("--timezone", default="")
    plan.add_argument("--attach-pid", action="store_true")
    plan.add_argument("--publish-mode", choices=["schedule", "draft"], default="schedule")
    plan.add_argument("--plan-name", default="schedule.generated.csv")

    for name in ("validate", "preview", "publish"):
        action = commands.add_parser(name)
        add_adspower_work_dir(action)
        action.add_argument("--input-file", default="")
        if name != "validate":
            action.add_argument("--visible", action="store_true")
        if name == "publish":
            action.add_argument("--publish-code", default="")

    add_adspower_work_dir(commands.add_parser("runtime"))
    return parser


def adspower_command(argv: list[str]) -> int:
    args = build_adspower_parser().parse_args(argv)
    runtime = ensure_adspower_runtime(install=True)
    node = Path(str(runtime["node"]))
    work_dir = Path(args.work_dir).expanduser().resolve()
    workspace = initialize_adspower_workspace(work_dir)
    command = args.ads_command
    if command == "runtime":
        print_json({"ready": True, "stage": "runtime", "runtime": runtime, "workspace": workspace})
        return 0
    if command == "init":
        print_json({"ready": True, "stage": "init", "runtime": runtime, "workspace": workspace, "next_state": "publish-source"})
        return 0
    if command == "plan":
        print_json(create_adspower_plan(args, work_dir))
        return 0
    if command == "validate":
        input_path = resolve_adspower_input(args, work_dir, workspace)
        tasks_path = work_dir / "tasks.validated.json"
        validation = prepare_adspower_tasks(input_path, tasks_path, "preview")
        print_json(
            {
                "ready": True,
                "stage": "validate",
                "input_file": str(input_path),
                "tasks": str(tasks_path),
                "count": validation["tasks"],
                "next_state": "publish-review",
            }
        )
        return 0
    if command == "profiles":
        output_path = work_dir / "profiles.json"
        result = run_adspower_node(
            node,
            [str(ADSPOWER_CLI), "--config", workspace["config"], "--list-windows", "--out", str(output_path)],
            work_dir=work_dir,
            timeout=180,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stdout[-2000:])
        profiles = load_json_value(output_path)
        print_json({"ready": True, "stage": "profiles", "profiles_file": str(output_path), "profiles": profiles})
        return 0
    if command == "check":
        output_path = work_dir / "logs" / f"check-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
        node_args = [
            str(ADSPOWER_CLI),
            "--config",
            workspace["config"],
            "--check-tiktok-upload",
            "--concurrency",
            str(max(1, args.concurrency)),
            "--out",
            str(output_path),
        ]
        if args.profile_no:
            node_args.extend(["--names", ",".join(parse_profiles(args.profile_no))])
        result = run_adspower_node(
            node,
            node_args,
            work_dir=work_dir,
            timeout=3600,
            env_overrides={"ADSPOWER_PUBLISH_HEADLESS": "0" if args.visible else "1"},
        )
        report = load_json_value(output_path)
        print_json(
            {
                "ready": result.returncode == 0,
                "stage": "check",
                "report": str(output_path),
                "results": report if isinstance(report, list) else [],
                "error_category": "" if result.returncode == 0 else adspower_error_category(result.stdout),
                "error": "" if result.returncode == 0 else safe_reason(result.stdout[-1500:]),
            }
        )
        return result.returncode
    return execute_adspower_plan(command, args, node, work_dir, workspace)


def install_home() -> Path:
    override = os.environ.get("H_INSTALL_HOME", "").strip()
    return Path(override).expanduser().resolve() if override else Path.home().resolve()


def merge_personal_marketplace(path: Path) -> None:
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"Personal marketplace JSON is invalid: {path}") from exc
        if not isinstance(data, dict):
            raise RuntimeError(f"Personal marketplace must contain a JSON object: {path}")
    else:
        data = {}
    data.setdefault("name", "personal")
    interface = data.setdefault("interface", {})
    if not isinstance(interface, dict):
        interface = {}
        data["interface"] = interface
    interface.setdefault("displayName", "Personal")
    plugins = data.get("plugins", [])
    if not isinstance(plugins, list):
        raise RuntimeError(f"Personal marketplace plugins must be a list: {path}")
    entry = {
        "name": "h",
        "source": {"source": "local", "path": "./plugins/h"},
        "policy": {"installation": "INSTALLED_BY_DEFAULT", "authentication": "ON_USE"},
        "category": "Productivity",
    }
    merged: list[object] = []
    replaced = False
    for plugin in plugins:
        if isinstance(plugin, dict) and plugin.get("name") == "h":
            if not replaced:
                merged.append(entry)
                replaced = True
            continue
        merged.append(plugin)
    if not replaced:
        merged.append(entry)
    data["plugins"] = merged
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        backup = path.with_name(f"marketplace.backup-{int(time.time())}.json")
        shutil.copy2(path, backup)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    temporary.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def install_local() -> int:
    packaged_core = packaged_core_path()
    if not packaged_core:
        raise RuntimeError("This package does not contain the portable H runtime.")
    source = PLUGIN_ROOT.resolve()
    home = install_home()
    marketplace_root = home / ".agents" / "plugins"
    target = marketplace_root / "plugins" / "h"
    target.parent.mkdir(parents=True, exist_ok=True)
    backup: Path | None = None
    if source != target.resolve():
        staging = target.parent / f".h.installing-{os.getpid()}-{int(time.time())}"
        ignore = shutil.ignore_patterns(
            ".git",
            ".github",
            ".agents",
            ".h_api_key",
            ".h_ready.json",
            ".h_venv",
            ".ruff_cache",
            "__pycache__",
            "*.pyc",
            "build",
            "dist",
            "portable-dist",
        )
        preserved_key = (target / ".h_api_key").read_bytes() if (target / ".h_api_key").is_file() else None
        shutil.copytree(source, staging, ignore=ignore)
        if preserved_key:
            (staging / ".h_api_key").write_bytes(preserved_key)
        if target.exists():
            backup = target.parent / f"h.backup-{int(time.time())}-{os.getpid()}"
            target.rename(backup)
        staging.rename(target)
    runtime_dir = target / "runtime"
    launcher_name = "h_launcher.exe" if os.name == "nt" else "h_launcher"
    core_name = "h_core.exe" if os.name == "nt" else "h_core"
    installed_launcher = runtime_dir / launcher_name
    installed_core = runtime_dir / core_name
    if not installed_launcher.is_file() or not installed_core.is_file():
        raise RuntimeError("The copied H package is missing its portable runtime files.")
    if os.name != "nt":
        installed_launcher.chmod(installed_launcher.stat().st_mode | 0o111)
        installed_core.chmod(installed_core.stat().st_mode | 0o111)
        if sys.platform == "darwin":
            run_quiet(["xattr", "-dr", "com.apple.quarantine", str(target)], timeout=30)
    environment = {"CODEX_HOME": str(home / ".codex")} if os.environ.get("H_INSTALL_HOME") else None
    smoke = run_quiet(
        [str(installed_launcher), "start", "--offline"],
        cwd=target,
        timeout=300,
        env_overrides=environment,
    )
    smoke_payload = parse_last_json(smoke.stdout)
    if smoke.returncode != 0 or not smoke_payload.get("ready"):
        if backup and target.exists():
            failed = target.parent / f"h.failed-{int(time.time())}-{os.getpid()}"
            target.rename(failed)
            backup.rename(target)
        raise RuntimeError("The installed H portable runtime did not pass its offline startup check.")
    marketplace_path = marketplace_root / "marketplace.json"
    merge_personal_marketplace(marketplace_path)
    print_json(
        {
            "ready": True,
            "state": "installed",
            "plugin_path": str(target),
            "marketplace_path": str(marketplace_path),
            "backup_path": str(backup) if backup else "",
            "portable_runtime": True,
            "display_text": (
                "H 安装完成，内置运行环境已通过检查，不需要安装 Python。\n"
                "请完全退出并重新打开 Codex，然后新建任务调用 H。"
            ),
        }
    )
    return 0


def set_key() -> int:
    value = clean_secret(getpass.getpass("Kie API key: "))
    if not value:
        raise RuntimeError("No Kie API key was entered.")
    if value == read_secret_file(FASTMOSS_KEY_FILE):
        raise RuntimeError("Kie API key cannot be identical to the configured FastMoss API key.")
    write_secret_file(USER_KEY_FILE, value)
    print(f"Kie API key saved to {USER_KEY_FILE}", flush=True)
    return 0


def set_fastmoss_key() -> int:
    value = clean_secret(os.environ.pop("H_FASTMOSS_KEY_INPUT", ""))
    if not value:
        value = clean_secret(getpass.getpass("FastMoss API key: "))
    if not value:
        raise RuntimeError("No FastMoss API key was entered.")
    if value in configured_kie_keys():
        raise RuntimeError("FastMoss API key cannot be identical to the configured Kie API key.")
    write_secret_file(FASTMOSS_KEY_FILE, value)
    print_json(
        {
            "ready": True,
            "state": "fastmoss-key-saved",
            "key_file": str(FASTMOSS_KEY_FILE),
            "display_text": "FastMoss API Key 已安全保存到当前用户目录，未写入插件或日志。",
        }
    )
    return 0


def startup_failure(exc: Exception) -> dict[str, Any]:
    reason = safe_reason(exc)
    return {
        "ready": False,
        "state": "setup-error",
        "error_category": "runtime",
        "protocol_version": PROTOCOL_VERSION,
        "display_text": (
            f"{GREETING}\n\nH 自动环境准备失败。\n归因：本地运行环境异常。\n"
            f"原因：{reason}\nH 尚未提交任何生成任务。"
        ),
    }


def main(argv: list[str]) -> int:
    configure_utf8_output()
    command = argv[0] if argv else "start"
    try:
        if command == "install-local":
            return install_local()
        if command == "set-key":
            return set_key()
        if command == "set-fastmoss-key":
            return set_fastmoss_key()
        if command in {"bootstrap", "--bootstrap"}:
            return doctor(offline=True)
        if command in {"--doctor", "doctor"}:
            extra = [value for value in argv[1:] if value != "--offline"]
            return doctor(offline="--offline" in argv[1:], forwarded_args=extra)
        if command == "start":
            capability = "menu"
            extra: list[str] = []
            index = 1
            while index < len(argv):
                value = argv[index]
                if value == "--capability":
                    if index + 1 >= len(argv):
                        raise ValueError("--capability requires pid, generate, publish, or menu")
                    capability = argv[index + 1]
                    index += 2
                    continue
                if value not in {"--offline", "--force-check"}:
                    extra.append(value)
                index += 1
            return start(
                offline="--offline" in argv[1:],
                force_check="--force-check" in argv[1:],
                capability=capability,
                forwarded_args=extra,
            )
        if command == "protocol":
            if len(argv) != 2:
                raise ValueError("Usage: h_run protocol <state>")
            return protocol(argv[1])
        if command == "adspower":
            report = bootstrap()
            if report.runtime_source != "packaged-executable" and Path(sys.executable).resolve() != Path(report.python).resolve():
                return subprocess.call(
                    [report.python, str(Path(__file__).resolve()), *argv],
                    cwd=str(PLUGIN_ROOT),
                    env={**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
                )
            return adspower_command(argv[1:])
        if command == "fastmoss":
            report = bootstrap()
            if report.runtime_source != "packaged-executable" and Path(sys.executable).resolve() != Path(report.python).resolve():
                return subprocess.call(
                    [report.python, str(Path(__file__).resolve()), *argv],
                    cwd=str(PLUGIN_ROOT),
                    env={**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
                )
            return fastmoss_command(argv[1:])
        report = bootstrap()
        return subprocess.call(core_command(report, argv), cwd=str(PLUGIN_ROOT))
    except Exception as exc:
        if command in {"start", "protocol"}:
            print_json(startup_failure(exc))
            return 0
        if command == "adspower":
            print_json(
                {
                    "ready": False,
                    "stage": "publish",
                    "error_category": adspower_error_category(exc),
                    "error": safe_reason(exc),
                }
            )
            return 1
        if command == "fastmoss":
            details = exc.as_dict() if isinstance(exc, FastMossError) else {
                "error_category": "runtime",
                "error": safe_reason(exc),
            }
            print_json(
                {
                    "ready": False,
                    "state": "fastmoss-error",
                    "stage": "fastmoss-product",
                    **details,
                    "display_text": fastmoss_error_display(exc),
                }
            )
            return 1
        raise


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except Exception as exc:
        print_json({"ready": False, "error_category": "runtime", "error": safe_reason(exc)})
        raise SystemExit(1)
